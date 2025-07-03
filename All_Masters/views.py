from rest_framework import generics
from .models import *
from .serializers import *
from rest_framework.filters import SearchFilter
from rest_framework_simplejwt.authentication import JWTAuthentication
from rest_framework.permissions import IsAuthenticated

# GstMaster views
class GSTMasterListCreate(generics.ListCreateAPIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]
    queryset = TaxDetails.objects.all()
    serializer_class = GSTMasterSerializer

class GSTMasterRetrieveUpdateDestroy(generics.RetrieveUpdateDestroyAPIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]
    queryset = TaxDetails.objects.all()
    serializer_class = GSTMasterSerializer

# TaxCode Views
class TaxCodeListCreate(generics.ListCreateAPIView):
    queryset = TaxCode.objects.all()
    serializer_class = TaxCodeSerializer

class TaxCodeRetrieveUpdateDestroy(generics.RetrieveUpdateDestroyAPIView):
    queryset = TaxCode.objects.all()
    serializer_class = TaxCodeSerializer

# Cust-Wise GST Master
class Cust_WiseListCreate(generics.ListCreateAPIView):
    queryset = Cut_Wise.objects.all()
    serializer_class = Cut_WiseSerializer

class Cust_WiseRetrieveUpdateDestroy(generics.RetrieveUpdateDestroyAPIView):
    queryset = Cut_Wise.objects.all()
    serializer_class = Cut_WiseSerializer


# Supplier_Customer Views 
class Supplier_CustomerListCreate(generics.ListCreateAPIView):
    queryset = Supplier_Customer.objects.all()
    serializer_class = Supplier_Customer_Serializer

class Supplier_CustomerRetrieveUpdateDestroy(generics.RetrieveUpdateDestroyAPIView):
    queryset = Supplier_Customer.objects.all()
    serializer_class = Supplier_Customer_Serializer


# Supplier_Customer_Buyer Views 
class Buyer_ListCreate(generics.ListCreateAPIView):
    queryset = Buyer_Contact.objects.all()
    serializer_class = Buyer_Contact_Serializer

class Buyer_RetrieveUpdateDestroy(generics.RetrieveUpdateDestroyAPIView):
    queryset = Buyer_Contact.objects.all()
    serializer_class = Buyer_Contact_Serializer


# Supplier_Customer_Banks Views 
class Bank_ListCreate(generics.ListCreateAPIView):
    queryset = Bank_Details.objects.all()
    serializer_class = Bank_Details_Serializer

class Bank_RetrieveUpdateDestroy(generics.RetrieveUpdateDestroyAPIView):
    queryset = Bank_Details.objects.all()
    serializer_class = Bank_Details_Serializer


# Supplier Masters Type:- Customer,Supplier,JobWork,CSJW
from rest_framework import viewsets
from .models import Customer, Supplier, JobWork, CSJW
from .serializers import CustomerSerializer,SupplierSerializer,JobWorkSerializer,CSJWSerializer

class CustomerViewSet(viewsets.ModelViewSet):
    queryset = Customer.objects.all()
    serializer_class = CustomerSerializer

class SupplierViewSet(viewsets.ModelViewSet):
    queryset = Supplier.objects.all()
    serializer_class = SupplierSerializer

class JobWorkViewSet(viewsets.ModelViewSet):
    queryset = JobWork.objects.all()
    serializer_class = JobWorkSerializer

class CSJWViewSet(viewsets.ModelViewSet):
    queryset = CSJW.objects.all()
    serializer_class = CSJWSerializer


# Supplier_Customer_Masters_General Type_Field_API
class Type_ListCreate(generics.ListCreateAPIView):
    queryset = Type_Model.objects.all()
    serializer_class = Type_Serializer

class Type_RetrieveUpdateDestroy(generics.RetrieveUpdateDestroyAPIView):
    queryset = Type_Model.objects.all()
    serializer_class = Type_Serializer

#Supplier_Customer_Masters_General Region_Field_API
class Region_ListCreate(generics.ListCreateAPIView):
    queryset = Region_Model.objects.all()
    serializer_class = Region_Serializer
    filter_backends = [SearchFilter]
    search_fields = ['RegionCode', 'RegionName']

class Region_RetrieveUpdateDestroy(generics.RetrieveUpdateDestroyAPIView):
    queryset = Region_Model.objects.all()
    serializer_class = Region_Serializer

# Supplier_Customer_Masters_General State_Field_API
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from .constants import INDIA_STATES_AND_UTS
from .serializers import StateUTSerializer

class StateUTListView(APIView):
    def get(self, request):
        search_query = request.query_params.get('search', '').strip().lower()

        # Filter the states and UTs based on the search query
        if search_query:
            filtered_states_uts = [
                item for item in INDIA_STATES_AND_UTS
                if search_query in item['name'].lower() or search_query in item['code'].lower()
            ]
        else:
            filtered_states_uts = INDIA_STATES_AND_UTS

        # Serialize the filtered data
        serializer = StateUTSerializer(filtered_states_uts, many=True)
        
        return Response(serializer.data, status=status.HTTP_200_OK)



# Fetch_Region
class Fetch_Region_ListCreate(generics.ListAPIView):
    queryset = Region_Model.objects.all()
    serializer_class = Fetch_Region_Serializer
   

# Supplier_Customer_Masters_General StateCode_Field_API
class StateCode_ListCreate(generics.ListCreateAPIView):
    queryset = StateCode_Model.objects.all()
    serializer_class = StateCode_Serializer
    filter_backends = [SearchFilter]
    search_fields = ['RegionCode', 'RegionName']

class StateCode_RetrieveUpdateDestroy(generics.RetrieveUpdateDestroyAPIView):
    queryset = StateCode_Model.objects.all()
    serializer_class = StateCode_Serializer

# Supplier_Customer_Masters_General Payment_Term_Field_API
class Payment_Term_ListCreate(generics.ListCreateAPIView):
    queryset = PaymentTerm_Model.objects.all()
    serializer_class = PaymentTerm_Serializer

class Payment_Term_RetrieveUpdateDestroy(generics.RetrieveUpdateDestroyAPIView):
    queryset = PaymentTerm_Model.objects.all()
    serializer_class = PaymentTerm_Serializer

# Supplier_Customer_Masters_General Country_Field_API
class Country_ListCreate(generics.ListCreateAPIView):
    queryset = Country_Model.objects.all()
    serializer_class = Country_Serializer

class Country_RetrieveUpdateDestroy(generics.RetrieveUpdateDestroyAPIView):
    queryset = Country_Model.objects.all()
    serializer_class = Country_Serializer


# Supplier Genral Page Country
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from .constants import COUNTRY_NAMES
from .serializers import CountrySerializer

class CountryListView(APIView):
    def get(self, request):
        search_query = request.query_params.get('search', '').strip().lower()

        # Filter the country list based on the search query
        if search_query:
            filtered_countries = [
                {"name": country} for country in COUNTRY_NAMES
                if search_query in country.lower()
            ]
        else:
            filtered_countries = [{"name": country} for country in COUNTRY_NAMES]

        # Serialize the filtered country names
        serializer = CountrySerializer(filtered_countries, many=True)
        
        return Response(serializer.data, status=status.HTTP_200_OK)



# Supplier_Customer_Masters_General Currency_Field_API
class Currency_ListCreate(generics.ListCreateAPIView):
    queryset = Currency_Model.objects.all()
    serializer_class = Currency_Serializer

class Currency_RetrieveUpdateDestroy(generics.RetrieveUpdateDestroyAPIView):
    queryset = Currency_Model.objects.all()
    serializer_class = Currency_Serializer

# Supplier Genral Page Currency
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from .constants import CURRENCY_CODES
from .serializers import CurrencySerializer

class CurrencyCodeListView(APIView):
    def get(self, request):
        search_query = request.query_params.get('search', '').strip().lower()

        # Filter the country list based on the search query
        if search_query:
            filtered_currency = [
                {"code": currency} for currency in CURRENCY_CODES
                if search_query in currency.lower()
            ]
        else:
            filtered_currency = [{"code": currency} for currency in CURRENCY_CODES]

        # Serialize the filtered country names
        serializer = CurrencySerializer(filtered_currency, many=True)
        
        return Response(serializer.data, status=status.HTTP_200_OK)



# Supplier_Customer_Masters_General City_Field_API
class City_ListCreate(generics.ListCreateAPIView):
    queryset = City_Model.objects.all()
    serializer_class = City_Serializer

class City_RetrieveUpdateDestroy(generics.RetrieveUpdateDestroyAPIView):
    queryset = City_Model.objects.all()
    serializer_class = City_Serializer


# Supplier_Customer_Masters_General Sector_Field_API
class Sector_ListCreate(generics.ListCreateAPIView):
    queryset = Sector_Model.objects.all()
    serializer_class = Sector_Serializer

class Sector_RetrieveUpdateDestroy(generics.RetrieveUpdateDestroyAPIView):
    queryset = Sector_Model.objects.all()
    serializer_class = Sector_Serializer

# Supplier_Customer_Masters_General Group_Field_API
class Group_ListCreate(generics.ListCreateAPIView):
    queryset = Group_Model.objects.all()
    serializer_class = Group_Serializer

class Group_RetrieveUpdateDestroy(generics.RetrieveUpdateDestroyAPIView):
    queryset = Group_Model.objects.all()
    serializer_class = Group_Serializer

# Supplier_Customer_Masters_General QMSC_Code_Field_API
class QMSC_Code_ListCreate(generics.ListCreateAPIView):
    queryset = QMSC_Code_Model.objects.all()
    serializer_class = QMSC_Code_Serializer

class QMSC_Code_RetrieveUpdateDestroy(generics.RetrieveUpdateDestroyAPIView):
    queryset = QMSC_Code_Model.objects.all()
    serializer_class = QMSC_Code_Serializer

# Supplier_Customer_Masters:- Fetch Supplier List
class Supplier_List_ListCreate(generics.ListCreateAPIView):
    queryset = Supplier_Customer.objects.all()
    serializer_class = Supplier_List_Serializer
    filter_backends = [SearchFilter]
    search_fields = ['Name', 'Code_No']


# Supplier Master Type:- Customer, Supplier, JobWork, CSJW
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from .models import SupplierType, CustomerType, JobworkType, CSJWType
from .serializers import SupplierTypeSerializer, CustomerTypeSerializer, JobworkTypeSerializer, CSJWTypeSerializer

class EntityView(APIView):
    def get(self, request, entity_type):
        prefix_mapping = {
            'supplier': 'S',
            'customer': 'C',
            'jobwork': 'J',
            'csjw': 'CSJW'
        }
        if entity_type not in prefix_mapping:
            return Response({"error": "Invalid entity type"}, status=status.HTTP_400_BAD_REQUEST)

        prefix = prefix_mapping[entity_type]
        model_mapping = {
            'supplier': SupplierType,
            'customer': CustomerType,
            'jobwork': JobworkType,
            'csjw': CSJWType
        }
        model = model_mapping[entity_type]

        last_entity = model.objects.order_by('-id').first()
        if last_entity:
            last_code = last_entity.code
            last_number = int(last_code[len(prefix):])
            new_number = last_number + 1
            new_code = f"{prefix}{new_number:04d}"
        else:
            new_code = f"{prefix}0001"

        return Response({"next_code": new_code}, status=status.HTTP_200_OK)

    def post(self, request, entity_type):
        if entity_type == 'supplier':
            name = request.data.get('name')
            code = request.data.get('code')
            if not name or not code:
                return Response({"error": "Name and Code fields are required"}, status=status.HTTP_400_BAD_REQUEST)
            entity = SupplierType(name=name, code=code)
            serializer = SupplierTypeSerializer(entity)
        
        elif entity_type == 'customer':
            name = request.data.get('name')
            code = request.data.get('code')
            if not name or not code:
                return Response({"error": "Name and Code fields are required"}, status=status.HTTP_400_BAD_REQUEST)
            entity = CustomerType(name=name, code=code)
            serializer = CustomerTypeSerializer(entity)
        
        elif entity_type == 'jobwork':
            name = request.data.get('name')
            code = request.data.get('code')
            if not name or not code:
                return Response({"error": "Name and Code fields are required"}, status=status.HTTP_400_BAD_REQUEST)
            entity = JobworkType(name=name, code=code)
            serializer = JobworkTypeSerializer(entity)
        
        elif entity_type == 'csjw':
            name = request.data.get('name')
            code = request.data.get('code')
            if not name or not code:
                return Response({"error": "Name and Code fields are required"}, status=status.HTTP_400_BAD_REQUEST)
            entity = CSJWType(name=name, code=code)
            serializer = CSJWTypeSerializer(entity)
        
        else:
            return Response({"error": "Invalid entity type"}, status=status.HTTP_400_BAD_REQUEST)

        entity.save()
        return Response(serializer.data, status=status.HTTP_201_CREATED)





# Supplier_Customer_Masters:- State, CodeNo, City
# from rest_framework.response import Response
# from rest_framework.decorators import api_view
# from .constants import STATES_AND_CITIES
# from .serializers import StateSerializer

# @api_view(['GET'])
# def get_state_data(request):
#     state_name = request.GET.get('state')
#     if state_name in STATES_AND_CITIES:
#         state_data = STATES_AND_CITIES[state_name]
#         return Response({
#             'state': state_name,
#             'code': state_data['code'],
#             'Gst_Code': state_data['Gst_Code'],
#             'cities': state_data['cities']
#         })
#     return Response({'error': 'State not found'}, status=404)

from rest_framework.response import Response
from rest_framework.decorators import api_view
from rest_framework import status
from .constants import STATES_AND_CITIES
from .serializers import StateSerializer

@api_view(['GET'])
def get_state_data(request):
    state_name = request.query_params.get('state')
    
    if state_name:
        state_data = STATES_AND_CITIES.get(state_name)
        if state_data:
            # Serialize data using StateSerializer
            serializer = StateSerializer(data={
                'name': state_name,
                'code': state_data['code'],
                'gst_code': state_data['Gst_Code'],
                'cities': state_data['cities']
            })
            if serializer.is_valid():
                return Response(serializer.data)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        return Response({'error': 'State not found'}, status=status.HTTP_404_NOT_FOUND)
    
    # Return all states data if 'state' parameter is missing or empty
    all_states = [
        {
            'name': name,
            'code': data['code'],
            'gst_code': data['Gst_Code'],
            'cities': data['cities']
        }
        for name, data in STATES_AND_CITIES.items()
    ]
    
    return Response(all_states)





# Supplier_Customer_Masters:- GST TAX CODE
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from .constants import GST_STATE_CODES

class TaxTypeView(APIView):
    def get(self, request, state_code):
        tax_type = GST_STATE_CODES.get(state_code)
        if tax_type:
            return Response({"state_code": state_code, "tax_type": tax_type}, status=status.HTTP_200_OK)
        return Response({"error": "Invalid GST state code"}, status=status.HTTP_400_BAD_REQUEST)





# Cross Reference: Item Cross Reference
class Item_Cross_Reference_ListCreate(generics.ListCreateAPIView):
    queryset = Item_Cross_Model.objects.all()
    serializer_class = Item_Cross_Serializer

class Item_Cross_Reference_RetrieveUpdateDestroy(generics.RetrieveUpdateDestroyAPIView):
    queryset = Item_Cross_Model.objects.all()
    serializer_class = Item_Cross_Serializer

# Cross Reference: Supplier Item
class Supplier_Item_Reference_ListCreate(generics.ListCreateAPIView):
    queryset = Supplier_Item_Model.objects.all()
    serializer_class = Supplier_Item_Serializer

class Supplier_Item_RetrieveUpdateDestroy(generics.RetrieveUpdateDestroyAPIView):
    queryset = Supplier_Item_Model.objects.all()
    serializer_class = Supplier_Item_Serializer


# Cross Reference: Customer_Item_VA_Rate
class Item_VA_Rate_ListCreate(generics.ListCreateAPIView):
    queryset = Item_VA_Rate_Model.objects.all()
    serializer_class = Item_VA_Rate_Serializer

class Item_VA_Rate_RetrieveUpdateDestroy(generics.RetrieveUpdateDestroyAPIView):
    queryset = Item_VA_Rate_Model.objects.all()
    serializer_class = Item_VA_Rate_Serializer

# Cross Reference: Supplier Name Fetch
class Cross_Ref_Supplier_ListCreate(generics.ListCreateAPIView):
    queryset = Supplier_Customer.objects.all()
    serializer_class = Cross_Ref_Supplier_List_Serializer
    filter_backends = [SearchFilter]
    search_fields = ['Name', 'Code_No']


# Item Cross Reference:- Fetch Item
class Item_Cross_Item_ListCreate(generics.ListCreateAPIView):
    queryset = ItemMaster.objects.all()
    serializer_class = Item_Cross_Item_Serializer
    filter_backends = [SearchFilter]
    search_fields = ['SE_Item', 'Name_Description']


# New Work Center Master: Add New
class Work_Center_ListCreate(generics.ListCreateAPIView):
    queryset = Work_Center_Model.objects.all()
    serializer_class = Work_Center_Serializer

class Work_Center_RetrieveUpdateDestroy(generics.RetrieveUpdateDestroyAPIView):
    queryset = Work_Center_Model.objects.all()
    serializer_class = Work_Center_Serializer

# New Work Center Master: Machine Group Work Center Type Group
class Machine_Group_Work_Center_ListCreate(generics.ListCreateAPIView):
    queryset = Machine_Work_Center_Group_Model.objects.all()
    serializer_class = Machine_Group_Work_Center_Serializer

class Machine_Group_Work_Center_RetrieveUpdateDestroy(generics.RetrieveUpdateDestroyAPIView):
    queryset = Machine_Work_Center_Group_Model.objects.all()
    serializer_class = Machine_Group_Work_Center_Serializer


# New Work Center Master: Work Center Type Group
class WorkCenterTypeGroup_ListCreate(generics.ListCreateAPIView):
    queryset = WorkCenterTypeGroup_Model.objects.all()
    serializer_class = WorkCenterTypeGroup_Serializer

class WorkCenterTypeGroup_RetrieveUpdateDestroy(generics.RetrieveUpdateDestroyAPIView):
    queryset = WorkCenterTypeGroup_Model.objects.all()
    serializer_class = WorkCenterTypeGroup_Serializer

# Cycle Time Master
class Cycle_Time_Master_ListCreate(generics.ListCreateAPIView):
    queryset = Cycle_Time_Master_Model.objects.all()
    serializer_class = Cycle_Time_Master_Serializer

class Cycle_Time_Master_RetrieveUpdateDestroy(generics.RetrieveUpdateDestroyAPIView):
    queryset =Cycle_Time_Master_Model.objects.all()
    serializer_class = Cycle_Time_Master_Serializer

# Shift Master
class Shift_Master_ListCreate(generics.ListCreateAPIView):
    queryset = Shift_Master_Model.objects.all()
    serializer_class = Shift_Master_Serializer

class Shift_Master_RetrieveUpdateDestroy(generics.RetrieveUpdateDestroyAPIView):
    queryset =Shift_Master_Model.objects.all()
    serializer_class = Shift_Master_Serializer


# Add_New_Operator_Supervisor_Master
class Add_New_Operator_ListCreate(generics.ListCreateAPIView):
    queryset = Add_New_Operator_Model.objects.all()
    serializer_class = Add_New_Operator_Serializer

class Add_New_Operator_RetrieveUpdateDestroy(generics.RetrieveUpdateDestroyAPIView):
    queryset =Add_New_Operator_Model.objects.all()
    serializer_class = Add_New_Operator_Serializer

# Add_New_Operator_Supervisor_Master: Department_Master
class Department_Master_ListCreate(generics.ListCreateAPIView):
    queryset = Department_Master_Model.objects.all()
    serializer_class = Department_Master_Serializer

class Department_Master_RetrieveUpdateDestroy(generics.RetrieveUpdateDestroyAPIView):
    queryset = Department_Master_Model.objects.all()
    serializer_class = Department_Master_Serializer

# Add_New_Operator_Supervisor_Master: Department_Category_Master
class Department_Category_Master_ListCreate(generics.ListCreateAPIView):
    queryset = Department_Category_Master_Model.objects.all()
    serializer_class = Department_Category_Master_Serializer

class Department_Category_Master_RetrieveUpdateDestroy(generics.RetrieveUpdateDestroyAPIView):
    queryset = Department_Category_Master_Model.objects.all()
    serializer_class = Department_Category_Master_Serializer



# Add_New_Operator_Supervisor_Master: Designation Api Field
class Designation_Api_ListCreate(generics.ListCreateAPIView):
    queryset = Designation_Api_Model.objects.all()
    serializer_class = Designation_Api_Serializer

class Designation_Api_RetrieveUpdateDestroy(generics.RetrieveUpdateDestroyAPIView):
    queryset = Designation_Api_Model.objects.all()
    serializer_class = Designation_Api_Serializer

# Add_New_Operator_Supervisor_Master: Contractor Api Field
class Contractor_Api_ListCreate(generics.ListCreateAPIView):
    queryset = Contractor_Api_Model.objects.all()
    serializer_class = Contractor_Api_Serializer

class Contractor_Api_RetrieveUpdateDestroy(generics.RetrieveUpdateDestroyAPIView):
    queryset = Contractor_Api_Model.objects.all()
    serializer_class = Contractor_Api_Serializer

# Unit Conversion Master
class Unit_Conversion_ListCreate(generics.ListCreateAPIView):
    queryset = Unit_Conversion_Model.objects.all()
    serializer_class = Unit_Conversion_Serializer

class Unit_Conversion_RetrieveUpdateDestroy(generics.RetrieveUpdateDestroyAPIView):
    queryset = Unit_Conversion_Model.objects.all()
    serializer_class = Unit_Conversion_Serializer


# Contractor Master
class Contractor_Master_ListCreate(generics.ListCreateAPIView):
    queryset = Contractor_Master_Model.objects.all()
    serializer_class = ContractorMaster_Serializer

class Contractor_Master_RetrieveUpdateDestroy(generics.RetrieveUpdateDestroyAPIView):
    queryset = Contractor_Master_Model.objects.all()
    serializer_class = ContractorMaster_Serializer


# Item Master Genral Page
class Item_Master_Genral_Page_ListCreate(generics.ListCreateAPIView):
    queryset = ItemMaster.objects.all()
    serializer_class = ItemMaster_Serializer

class Item_Master_Genral_Page_RetrieveUpdateDestroy(generics.RetrieveUpdateDestroyAPIView):
    queryset = ItemMaster.objects.all()
    serializer_class = ItemMaster_Serializer


# Item Master models Data2
class Item_Master_Data2_ListCreate(generics.ListCreateAPIView):
    queryset = Item_Master_Data2_Model.objects.all()
    serializer_class = Item_Master_Data2_Serializer

class Item_Master_Data2_RetrieveUpdateDestroy(generics.RetrieveUpdateDestroyAPIView):
    queryset = Item_Master_Data2_Model.objects.all()
    serializer_class = Item_Master_Data2_Serializer

# Item Master Technical Specification
class Technical_Specification_ListCreate(generics.ListCreateAPIView):
    queryset = Technical_Specification_Model.objects.all()
    serializer_class = Technical_Specification_Serializer

class Technical_Specification_RetrieveUpdateDestroy(generics.RetrieveUpdateDestroyAPIView):
    queryset = Technical_Specification_Model.objects.all()
    serializer_class = Technical_Specification_Serializer

# Item Master NPD
class NPD_ListCreate(generics.ListCreateAPIView):
    queryset = NPD_Model.objects.all()
    serializer_class = NPD_Serializer

class NPD_RetrieveUpdateDestroy(generics.RetrieveUpdateDestroyAPIView):
    queryset = NPD_Model.objects.all()
    serializer_class = NPD_Serializer

# Item Master General Page Fields MainGroup
class MainGroup_ListCreate(generics.ListCreateAPIView):
    queryset = MainGroup_Model.objects.all()
    serializer_class = MainGroup_Serializer

class MainGroup_RetrieveUpdateDestroy(generics.RetrieveUpdateDestroyAPIView):
    queryset = MainGroup_Model.objects.all()
    serializer_class = MainGroup_Serializer

# Item Master General Page Fields UnitCode
class UnitCode_ListCreate(generics.ListCreateAPIView):
    queryset = UnitCode_Model.objects.all()
    serializer_class = UnitCode_Serializer

class UnitCode_RetrieveUpdateDestroy(generics.RetrieveUpdateDestroyAPIView):
    queryset = UnitCode_Model.objects.all()
    serializer_class = UnitCode_Serializer

# Item Master General Page Fields TDC
class TDC_ListCreate(generics.ListCreateAPIView):
    queryset = TDC_Model.objects.all()
    serializer_class = TDC_Serializer

class TDC_RetrieveUpdateDestroy(generics.RetrieveUpdateDestroyAPIView):
    queryset = TDC_Model.objects.all()
    serializer_class = TDC_Serializer

# Item Master General Page Fields MetalType
class MetalType_ListCreate(generics.ListCreateAPIView):
    queryset = MetalType_Model.objects.all()
    serializer_class = MetalType_Serializer

class MetalType_RetrieveUpdateDestroy(generics.RetrieveUpdateDestroyAPIView):
    queryset = MetalType_Model.objects.all()
    serializer_class = MetalType_Serializer

# Item Master General Page Fields Item Group
class Item_Group_ListCreate(generics.ListCreateAPIView):
    queryset = Item_Group_Item_Master_Model.objects.all()
    serializer_class = Item_Group_Serializer

class Item_Group_RetrieveUpdateDestroy(generics.RetrieveUpdateDestroyAPIView):
    queryset = Item_Group_Item_Master_Model.objects.all()
    serializer_class = Item_Group_Serializer

# Item Master General Page Fields Store Location
class Store_Location_ListCreate(generics.ListCreateAPIView):
    queryset = Store_Location_Item_Master_Model.objects.all()
    serializer_class = Store_Location_Serializer

class Store_Location_RetrieveUpdateDestroy(generics.RetrieveUpdateDestroyAPIView):
    queryset = Store_Location_Item_Master_Model.objects.all()
    serializer_class = Store_Location_Serializer

# Item Master General Page Fields Route
class Route_ListCreate(generics.ListCreateAPIView):
    queryset = Route_Item_Master_Model.objects.all()
    serializer_class = Route_Serializer

class Route_RetrieveUpdateDestroy(generics.RetrieveUpdateDestroyAPIView):
    queryset = Route_Item_Master_Model.objects.all()
    serializer_class = Route_Serializer

# Item Master General Page Fields Parent FG Code
class Parent_FG_Code_ListCreate(generics.ListCreateAPIView):
    queryset = Parent_FG_Code_Item_Master_Model.objects.all()
    serializer_class = Parent_FG_Code_Serializer

class Parent_FG_Code_RetrieveUpdateDestroy(generics.RetrieveUpdateDestroyAPIView):
    queryset = Parent_FG_Code_Item_Master_Model.objects.all()
    serializer_class = Parent_FG_Code_Serializer

# Item Master General Page Fields Sector
class Sector_Item_master_ListCreate(generics.ListCreateAPIView):
    queryset = Sector_Item_Master_Model.objects.all()
    serializer_class = Sector_Item_Master_Serializer

class Sector_Item_master_RetrieveUpdateDestroy(generics.RetrieveUpdateDestroyAPIView):
    queryset = Sector_Item_Master_Model.objects.all()
    serializer_class = Sector_Item_Master_Serializer

# Item Master General Page Fields Item Section
class Item_Section_ListCreate(generics.ListCreateAPIView):
    queryset = Item_Section_Item_Master_Model.objects.all()
    serializer_class = Item_Section_Serializer

class Item_Section_RetrieveUpdateDestroy(generics.RetrieveUpdateDestroyAPIView):
    queryset = Item_Section_Item_Master_Model.objects.all()
    serializer_class = Item_Section_Serializer

# Item Master General Page Fields Type/Grade New
class Type_Grade_New_ListCreate(generics.ListCreateAPIView):
    queryset = Type_Grade_New_Model.objects.all()
    serializer_class = Type_Grade_New_Serializer

class Type_Grade_New_RetrieveUpdateDestroy(generics.RetrieveUpdateDestroyAPIView):
    queryset = Type_Grade_New_Model.objects.all()
    serializer_class = Type_Grade_New_Serializer

# Item Master General Page Fields Sub Group
class Sub_Group_ListCreate(generics.ListCreateAPIView):
    queryset = Sub_Group_Item_Master_Model.objects.all()
    serializer_class = Sub_Group_Serializer

class Sub_Group_RetrieveUpdateDestroy(generics.RetrieveUpdateDestroyAPIView):
    queryset = Sub_Group_Item_Master_Model.objects.all()
    serializer_class = Sub_Group_Serializer

# Item Master Data2 Page Fields Qty_Packing
class Qty_Packing_ListCreate(generics.ListCreateAPIView):
    queryset = Qty_Packing_Model.objects.all()
    serializer_class = Qty_Packing_Serializer

class Qty_Packing_RetrieveUpdateDestroy(generics.RetrieveUpdateDestroyAPIView):
    queryset = Qty_Packing_Model.objects.all()
    serializer_class = Qty_Packing_Serializer


# Item Masters:- Fetch Item List
class ItemMaster_List_ListCreate(generics.ListCreateAPIView):
    queryset = ItemMaster.objects.all()
    serializer_class = ItemMaster_List_Serializer
    filter_backends = [SearchFilter]
    search_fields = ['SE_Item', 'Name_Description']


# Item Master UnitCode HardCoded
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from .constants import UNIT_CODE 
from .serializers import UnitCodeHardCodedSerializer

class UnitCodeHardCodeListView(APIView):
    def get(self, request):
        search_query = request.query_params.get('search', '').strip().lower()

        # Filter the country list based on the search query
        if search_query:
            filtered_UnitCode = [
                {"name": UnitCode} for UnitCode in UNIT_CODE
                if search_query in UnitCode.lower()
            ]
        else:
            filtered_UnitCode = [{"name": UnitCode} for UnitCode in UNIT_CODE]

        # Serialize the filtered country names
        serializer = UnitCodeHardCodedSerializer(filtered_UnitCode, many=True)
        
        return Response(serializer.data, status=status.HTTP_200_OK)


# Cost Center Master
class Cost_Center_Master_ListCreate(generics.ListCreateAPIView):
    queryset = Cost_Center_Model.objects.all()
    serializer_class = Cost_Center_Master_Serializer

class Cost_Center_Master_RetrieveUpdateDestroy(generics.RetrieveUpdateDestroyAPIView):
    queryset = Cost_Center_Model.objects.all()
    serializer_class = Cost_Center_Master_Serializer

# Cost Center Master Add New
class Cost_Center_Master_Add_New_ListCreate(generics.ListCreateAPIView):
    queryset = Cost_Center_Master_Add_New_Model.objects.all()
    serializer_class = Cost_Center_Master_Add_New_Serializer

class Cost_Center_Master_Add_New_RetrieveUpdateDestroy(generics.RetrieveUpdateDestroyAPIView):
    queryset = Cost_Center_Master_Add_New_Model.objects.all()
    serializer_class = Cost_Center_Master_Add_New_Serializer

# Price List: Price List Master - Add Price List
class Add_Price_List_ListCreate(generics.ListCreateAPIView):
    queryset = Add_Price_List_Model.objects.all()
    serializer_class = Add_Price_List_Serializer

class Add_Price_List_RetrieveUpdateDestroy(generics.RetrieveUpdateDestroyAPIView):
    queryset = Add_Price_List_Model.objects.all()
    serializer_class = Add_Price_List_Serializer

# Price List: Price List Entry
class Price_List_Entry_ListCreate(generics.ListCreateAPIView):
    queryset = Price_List_Entry_Model.objects.all()
    serializer_class = Price_List_Entry_Serializer

class Price_List_Entry_RetrieveUpdateDestroy(generics.RetrieveUpdateDestroyAPIView):
    queryset = Price_List_Entry_Model.objects.all()
    serializer_class = Price_List_Entry_Serializer

# Bom Routing Master: - Production Department Master
class Production_Department_ListCreate(generics.ListCreateAPIView):
    queryset = Production_Department_Model.objects.all()
    serializer_class = Production_Department_Serializer

class Production_Department_RetrieveUpdateDestroy(generics.RetrieveUpdateDestroyAPIView):
    queryset = Production_Department_Model.objects.all()
    serializer_class = Production_Department_Serializer

# Bom Routing Master: - Operation Master
class Operation_Master_ListCreate(generics.ListCreateAPIView):
    queryset = Operation_Master_Model.objects.all()
    serializer_class = Operation_Master_Serializer

class Operation_Master_RetrieveUpdateDestroy(generics.RetrieveUpdateDestroyAPIView):
    queryset = Operation_Master_Model.objects.all()
    serializer_class = Operation_Master_Serializer

# Bom Routing Master: - Std Routing
class Std_Routing_ListCreate(generics.ListCreateAPIView):
    queryset = Std_Routing_Model.objects.all()
    serializer_class = Std_Routing_Serializer

class Std_Routing_RetrieveUpdateDestroy(generics.RetrieveUpdateDestroyAPIView):
    queryset = Std_Routing_Model.objects.all()
    serializer_class = Std_Routing_Serializer

# Bom Routing Master: - Bom Item Group Details
class Bom_Item_Group_ListCreate(generics.ListCreateAPIView):
    queryset = Bom_Item_Group_Model.objects.all()
    serializer_class = Bom_Item_Group_Serializer

class Bom_Item_Group_RetrieveUpdateDestroy(generics.RetrieveUpdateDestroyAPIView):
    queryset = Bom_Item_Group_Model.objects.all()
    serializer_class = Bom_Item_Group_Serializer
















#GST RATE EXCEL UPLOAD
from rest_framework import status
from rest_framework.response import Response
from rest_framework.views import APIView
from django.http import HttpResponse
from .models import ExcelFile
from .serializers import ExcelFileSerializer

class ExcelFileUploadView(APIView):
    def post(self, request, *args, **kwargs):
        file = request.FILES.get('file')
        if not file:
            return Response({"error": "No file provided"}, status=status.HTTP_400_BAD_REQUEST)

        instance = ExcelFile.objects.create(file=file)
        serializer = ExcelFileSerializer(instance)
        return Response(serializer.data, status=status.HTTP_201_CREATED)

class ExcelFileDownloadView(APIView):
    def get(self, request, *args, **kwargs):
        file_id = kwargs.get('pk')
        try:
            excel_file = ExcelFile.objects.get(pk=file_id)
        except ExcelFile.DoesNotExist:
            return Response({"error": "File not found"}, status=status.HTTP_404_NOT_FOUND)

        file_path = excel_file.file.path
        with open(file_path, 'rb') as f:
            response = HttpResponse(f.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="{excel_file.file.name}"'
            return response


# Item Main Group Api
from rest_framework import viewsets
from .models import Tool
from .serializers import ToolSerializer

class ToolViewSet(viewsets.ModelViewSet):
    queryset = Tool.objects.all()
    serializer_class = ToolSerializer



########GST NUMBER
from rest_framework import status
from rest_framework.response import Response
from rest_framework.views import APIView

class GSTNumberInfoView(APIView):
    def get(self, request, *args, **kwargs):
        # Get query parameters
        state_code = request.GET.get('state_code')
        pan_number = request.GET.get('pan_number')

        if not state_code or not pan_number:
            return Response(
                {"error": "state_code and pan_number are required."},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Validate input
        if not state_code.isdigit() or len(state_code) != 2:
            return Response(
                {"error": "state_code must be a 2-digit numeric value."},
                status=status.HTTP_400_BAD_REQUEST
            )
        if not pan_number.isalnum() or len(pan_number) != 10:
            return Response(
                {"error": "pan_number must be a 10 alphanumeric characters."},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Combine the state_code and pan_number
        gst_number = f"{state_code}{pan_number}"

        # Return the combined GST number
        return Response(
            {"gst_number": gst_number},
            status=status.HTTP_200_OK
        )

# Supplier Sub Module

class SupplierViewSet(viewsets.ModelViewSet):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]
    queryset = Item.objects.all()
    serializer_class = ItemSerializer

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)


from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.generics import ListAPIView
from django.shortcuts import get_object_or_404
from django.template.loader import get_template
from weasyprint import HTML
from django.http import HttpResponse
from .models import Item
from .serializers import ItemPDFSerializer


# 1. PDF Data API (returns JSON with required fields only)
class ItemPDFDataAPIView(APIView):
    def get(self, request, pk):
        item = get_object_or_404(Item, pk=pk)
        serializer = ItemPDFSerializer(item)
        return Response(serializer.data)


# 2. PDF File Generator (renders Item + Buyer_Contact + Bank_Details into one PDF)
def generate_item_pdf(request, pk):
    item = get_object_or_404(Item, pk=pk)
    buyer_contacts = item.buyer_contacts.all()
    bank_details = item.bank_details.all()

    template = get_template('SupplierReport.html')  # Make sure this template exists
    html = template.render({
        'item': item,
        'buyer_contacts': buyer_contacts,
        'bank_details': bank_details
    })

    pdf_file = HTML(string=html).write_pdf()
    response = HttpResponse(pdf_file, content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="item_{pk}.pdf"'
    return response


from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework_simplejwt.authentication import JWTAuthentication
from .models import Item
from django.db.models import Q

class ItemSummaryAPIView(APIView):
     

    def get(self, request, *args, **kwargs):
        # Get query parameters
        item_type = request.query_params.get('type')
        name = request.query_params.get('name')
        number = request.query_params.get('number')

        # Start with all items
        items = Item.objects.all()

        # Apply filters based on query parameters
        if item_type:
            items = items.filter(type__icontains=item_type)
        if name:
            items = items.filter(Name__icontains=name)
        if number:
            items = items.filter(number__icontains=number)

        # Prepare response data
        data = [
            {
                "id": item.id,
                "type": item.type,
                "number": item.number,
                "Name": item.Name,
                "Contact_No": item.Contact_No,
                "Email_Id": item.Email_Id,
                "Vendor_Code": item.Vendor_Code,
                "Payment_Term": item.Payment_Term,
                "GST_No": item.GST_No,
                "GST_No2": item.GST_No2,
                "GST_Tax_Code": item.GST_Tax_Code,
                "Auth": item.is_verified,
                "User": item.created_by.username if item.created_by else None,
                "View": f"/All_Masters/api/item/pdf/{item.id}/",
                "Edit": f"/All_Masters/api/SupplierData/{item.id}/"
            }
            for item in items
        ]
        return Response(data)


from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from .models import Item
import re

class GetNextItemNumber(APIView):
    def get(self, request, *args, **kwargs):
        item_type = request.GET.get('type', None)

        if not item_type:
            return Response({"error": "type is required"}, status=status.HTTP_400_BAD_REQUEST)

        prefix_map = {
            "Customer": "C",
            "Supplier": "S",
            "Job Work": "JW",
            "C/S/J/W": "CSJW"
        }

        prefix = prefix_map.get(item_type)
        if not prefix:
            return Response({"error": "Invalid type provided"}, status=status.HTTP_400_BAD_REQUEST)

        # Filter items by type
        filtered_items = Item.objects.filter(type=item_type)

        # Get all valid codes for the type that match the prefix
        max_number = 0
        for item in filtered_items:
            code = item.number or ""
            match = re.match(rf"^{prefix}(\d+)$", code)
            if match:
                num = int(match.group(1))
                if num > max_number:
                    max_number = num

        # Calculate next number
        next_num = max_number + 1
        formatted_next_number = f"{prefix}{str(next_num).zfill(3)}"

        return Response({"next_number": formatted_next_number})



# Supplier Sub Module Excel Generate Code
import openpyxl
from openpyxl.styles import Font, Alignment
from django.http import HttpResponse
from .models import Item

def export_items_to_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Items"

    # Define fields
    fields = [field.name for field in Item._meta.fields if field.name != 'created_by']
    fields.append('created_by')

    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(fields))
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = 'Items Export'
    title_cell.font = Font(size=14, bold=True)
    title_cell.alignment = Alignment(horizontal='center')

    # Header row
    header_font = Font(bold=True)
    for col_num, column_title in enumerate(fields, 1):
        cell = ws.cell(row=2, column=col_num, value=column_title)
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Data rows
    for row_num, item in enumerate(Item.objects.all(), start=3):
        for col_num, field in enumerate(fields, 1):
            value = getattr(item, field)
            if field == 'created_by':
                value = value.username if value else ''
            ws.cell(row=row_num, column=col_num, value=value)

    # Auto-adjust column widths (skip merged cells)
    from openpyxl.utils import get_column_letter
    for col in range(1, len(fields) + 1):
        column_letter = get_column_letter(col)
        max_length = 0
        for row in ws.iter_rows(min_row=2, min_col=col, max_col=col):
            for cell in row:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
        ws.column_dimensions[column_letter].width = max_length + 4

    # Return Excel file in response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=items_only_export.xlsx'
    wb.save(response)
    return response


################################



################################


################################


################################


from rest_framework import viewsets
from .models import BOMItem
from .serializers import BOMItemSerializer

class ItemViewSet(viewsets.ModelViewSet):
    queryset = BOMItem.objects.all()
    serializer_class = BOMItemSerializer



#Bom Routing & Bill of Material(BOM):- Fetch Scrap Code



#BOM: Item Part Master
from .models import BOM_ItemPartMaster
from .serializers import BOM_ItemPartMaster_Serializer

# class BOM_ItemPartMaster_ListCreate(generics.ListCreateAPIView):
#     queryset = BOM_ItemPartMaster.objects.all()
#     serializer_class = BOM_ItemPartMaster_Serializer

# class BOM_ItemPartMaster_RetrieveUpdateDestroy(generics.RetrieveUpdateDestroyAPIView):
#     queryset = BOM_ItemPartMaster.objects.all()
#     serializer_class = BOM_ItemPartMaster_Serializer
# views.py
from rest_framework import generics
from .models import BOM_ItemPartMaster
from .serializers import BOM_ItemPartMaster_Serializer

class BOM_ItemPartMaster_ListCreate(generics.ListCreateAPIView):
    serializer_class = BOM_ItemPartMaster_Serializer

    def get_queryset(self):
        item_id = self.kwargs.get('item_id')
        return BOM_ItemPartMaster.objects.filter(item_id=item_id)

    def perform_create(self, serializer):
        item_id = self.kwargs.get('item_id')
        item_instance = ItemTable.objects.get(id=item_id)
        serializer.save(item=item_instance)

class BOM_ItemPartMaster_RetrieveUpdateDestroy(generics.RetrieveUpdateDestroyAPIView):
    serializer_class = BOM_ItemPartMaster_Serializer

    def get_queryset(self):
        item_id = self.kwargs.get('item_id')
        return BOM_ItemPartMaster.objects.filter(item_id=item_id)


from rest_framework.views import APIView
from rest_framework.response import Response
from .models import Operation_Master_Model
from .serializers import OperationMasterSerializer

class OperationDropdownView(APIView):
    def get(self, request):
        operations = Operation_Master_Model.objects.all().values('Operation_Name', 'Prefix').distinct()
        return Response(operations)

    


class PartViewSet(viewsets.ViewSet):
    def list(self, request):
        parts = PartCode.objects.all()
        serializer = PartSerializer(parts, many=True)
        return Response(serializer.data)
    
    def retrieve(self, request, pk=None):
        try:
            part = PartCode.objects.get(name=pk)
            serializer = PartSerializer(part)
            return Response(serializer.data)
        except PartCode.DoesNotExist:
            return Response({'error': 'Part not found'}, status=404)
        

# Item Master
from rest_framework import viewsets
from rest_framework.permissions import IsAuthenticated
from rest_framework_simplejwt.authentication import JWTAuthentication
from .models import ItemTable
from .serializers import ItemTableSerializer

class ItemTableViewSet(viewsets.ModelViewSet):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]
    queryset = ItemTable.objects.all()
    serializer_class = ItemTableSerializer

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)

        

from django.shortcuts import get_object_or_404
from django.http import HttpResponse
from django.template.loader import get_template
from weasyprint import HTML
from .models import ItemTable

def generate_ItemMaster_pdf(request, pk):
    item = get_object_or_404(ItemTable, pk=pk)
    item_master_data = getattr(item, 'item_master_data', None)
    technical_specs = item.technical_specifications.all()
    npd_details = item.npd_details.all()

    template = get_template('ItemMaster.html')  # You must create this template
    html = template.render({
        'item': item,
        'item_master_data': item_master_data,
        'technical_specs': technical_specs,
        'npd_details': npd_details
    })

    pdf_file = HTML(string=html).write_pdf()
    response = HttpResponse(pdf_file, content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="item_{pk}.pdf"'
    return response

from rest_framework.views import APIView
from rest_framework.response import Response
from django.db.models import Q
from .models import ItemTable

class ItemMasterAPIView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        queryset = ItemTable.objects.all()

        # Search fields
        item_search = request.GET.get("ItemSearch", "")
        if item_search:
            queryset = queryset.filter(
                Q(part_no__icontains=item_search) |
                Q(Name_Description__icontains=item_search) |
                Q(Part_Code__icontains=item_search)
            )

        # Filter by main_group
        main_group = request.GET.get("main_group")
        if main_group:
            queryset = queryset.filter(main_group=main_group)

        # Filter by item_group
        item_group = request.GET.get("item_group")
        if item_group:
            queryset = queryset.filter(item_group=item_group)

        # Build response data
        data = []
        for item in queryset:
            data.append({
                "id": item.id,
                "part_no": item.part_no,
                "Name_Description": item.Name_Description,
                "Part_Code": item.Part_Code,
                "Item_Size": item.Item_Size,
                "main_group": item.main_group,
                "item_group": item.item_group,
                "Store_Location": item.Store_Location,
                "Unit_Code": item.Unit_Code,
                "HSN_SAC_Code": item.HSN_SAC_Code,
                "Auth": item.is_verified,
                "User": item.created_by.username if item.created_by else None,
                "View": f"/All_Masters/ItemMaster/pdf/{item.id}/",
                "Edit": f"/All_Masters/api/item-table/{item.id}/",
            })

        return Response(data)


from rest_framework.views import APIView
from rest_framework.response import Response
from .models import ItemTable

class ItemDetailAPIView(APIView):
    def get(self, request, pk, *args, **kwargs):
        item = get_object_or_404(ItemTable, pk=pk)
        item_master = getattr(item, 'item_master_data', None)
        technical_specs = item.technical_specifications.all()
        npd_details = item.npd_details.all()

        return Response({
            "item": {
                "main_group": item.main_group,
                "item_group": item.item_group,
                "part_no": item.part_no,
                "Rate": item.Rate,
                "Length": item.Length,
                "Store_Location": item.Store_Location,
            },
            "item_master_data": {
                field.name: getattr(item_master, field.name)
                for field in item_master._meta.fields if field.name != "id"
            } if item_master else {},
            "technical_specifications": [
                {
                    "Specification": spec.Specification,
                    "Parameter": spec.Parameter
                } for spec in technical_specs
            ],
            "npd_details": [
                {
                    "NPD": npd.NPD,
                    "CustomerName": npd.CustomerName,
                    "NPD_Qty": npd.NPD_Qty,
                    "NPD_Due_Date": npd.NPD_Due_Date
                } for npd in npd_details
            ]
        })

# Item Master MainGrop+ItemGroup Code Generated Code

from rest_framework import viewsets
from rest_framework.decorators import api_view
from rest_framework.response import Response
from .models import MainGroup2, ItemGroup2
from .serializers import MainGroupSerializer2, ItemGroupSerializer2

class MainGroupViewSet2(viewsets.ModelViewSet):
    queryset = MainGroup2.objects.all()
    serializer_class = MainGroupSerializer2


class ItemGroupViewSet2(viewsets.ModelViewSet):
    queryset = ItemGroup2.objects.all()
    serializer_class = ItemGroupSerializer2

from All_Masters.models import ItemTable as ItemTable2

class ItemTableViewSet2(viewsets.ModelViewSet):
    queryset = ItemTable2.objects.all()
    serializer_class = ItemTableSerializer2

from rest_framework.decorators import api_view
from rest_framework.response import Response
from .models import MainGroup2, ItemGroup2, ItemTable

@api_view(['GET'])
def generate_code(request):
    subgroup_name = request.GET.get('subgroup_name', '').strip()
    group_name = request.GET.get('group_name', '').strip()

    if not subgroup_name or not group_name:
        return Response({"error": "subgroup_name and group_name are required"}, status=400)

    # Get MainGroup by subgroup_name
    try:
        main_group = MainGroup2.objects.get(subgroup_name__iexact=subgroup_name)
        main_prefix = main_group.prefix.upper()
    except MainGroup2.DoesNotExist:
        return Response({"error": "Invalid subgroup_name"}, status=400)

    # Get ItemGroup by group_name
    try:
        item_group = ItemGroup2.objects.get(group_name__iexact=group_name)
        item_prefix = item_group.prefix.upper() if item_group.prefix else ''
    except ItemGroup2.DoesNotExist:
        return Response({"error": "Invalid group_name"}, status=400)

    # Build full prefix for code generation
    full_prefix = f"{main_prefix}{item_prefix}"

    # Fetch all existing part_no values that start with this prefix
    existing_part_nos = ItemTable2.objects.filter(part_no__startswith=full_prefix).values_list('part_no', flat=True)

    # Extract numeric parts
    used_numbers = set()
    for part_no in existing_part_nos:
        suffix = part_no.replace(full_prefix, "")
        if suffix.isdigit():
            used_numbers.add(int(suffix))

    # Find the next unused number starting from 1001
    new_number = 1001
    while new_number in used_numbers:
        new_number += 1

    new_code = f"{full_prefix}{new_number}"
    return Response({"generated_code": new_code})


#################################################### BOM


from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from .models import BOMItem, ItemTable
from .serializers import BOMItemSerializer

# 1. Create BOM items for a specific item using item_id
class CreateBOMAPIView(APIView):
    def post(self, request, item_id):
        try:
            item = ItemTable.objects.get(id=item_id)
        except ItemTable.DoesNotExist:
            return Response({'error': 'Item not found'}, status=404)

        bom_data = request.data.get('boms', [])
        if not isinstance(bom_data, list):
            return Response({'error': 'Expected "boms" to be a list'}, status=400)

        for bom in bom_data:
            bom['item'] = item.id  # Assign the foreign key (item_id)

        serializer = BOMItemSerializer(data=bom_data, many=True)
        if serializer.is_valid():
            serializer.save()
            return Response({'message': 'BOM items created successfully'}, status=201)
        return Response(serializer.errors, status=400)


# 2. List all BOMs for an item using item_id
class ListBOMAPIView(APIView):
    def get(self, request, item_id):
        try:
            item = ItemTable.objects.get(id=item_id)
        except ItemTable.DoesNotExist:
            return Response({'error': 'Item not found'}, status=404)

        bom_items = BOMItem.objects.filter(item_id=item.id)
        serializer = BOMItemSerializer(bom_items, many=True)
        return Response(serializer.data)


# 3. Update a specific BOM item using item_id and bom_id
class UpdateBOMAPIView(APIView):
    def put(self, request, item_id, bom_id):
        try:
            item = ItemTable.objects.get(id=item_id)
        except ItemTable.DoesNotExist:
            return Response({'error': 'Item not found'}, status=404)

        try:
            bom_item = BOMItem.objects.get(id=bom_id, item_id=item.id)
        except BOMItem.DoesNotExist:
            return Response({'error': 'BOM item not found'}, status=404)

        serializer = BOMItemSerializer(bom_item, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response({'message': 'BOM item updated successfully'}, status=200)
        return Response(serializer.errors, status=400)


# 4. Delete a specific BOM item using item_id and bom_id
class DeleteBOMAPIView(APIView):
    def delete(self, request, item_id, bom_id):
        try:
            item = ItemTable.objects.get(id=item_id)
        except ItemTable.DoesNotExist:
            return Response({'error': 'Item not found'}, status=404)

        try:
            bom_item = BOMItem.objects.get(id=bom_id, item_id=item.id)
            bom_item.delete()
            return Response({'message': 'BOM item deleted successfully'}, status=200)
        except BOMItem.DoesNotExist:
            return Response({'error': 'BOM item not found'}, status=404)

# views.py

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from .models import BOMItem, ItemTable
from .serializers import BOMItemSerializer

class BOMListCreateAPIView(APIView):
    def get(self, request, item_id):
        try:
            item = ItemTable.objects.get(id=item_id)
        except ItemTable.DoesNotExist:
            return Response({'error': 'Item not found'}, status=404)

        boms = BOMItem.objects.filter(item=item)
        serializer = BOMItemSerializer(boms, many=True)
        return Response(serializer.data)

    def post(self, request, item_id):
        try:
            item = ItemTable.objects.get(id=item_id)
        except ItemTable.DoesNotExist:
            return Response({'error': 'Item not found'}, status=404)

        bom_data = request.data.copy()
        bom_data['item'] = item.id

        # Extract Operation and PartCode
        if 'PartCode' in bom_data and '|' in bom_data['PartCode']:
            try:
                operation_part, actual_partcode = [x.strip() for x in bom_data['PartCode'].split('|', 1)]
                bom_data['Operation'] = operation_part
                bom_data['PartCode'] = actual_partcode
            except ValueError:
                return Response({'error': 'Invalid PartCode format. Use: "OPERATION | PARTCODE"'}, status=400)

        # Split BomPartCode into part_no, part_code, description (but only store part_no)
        if 'BomPartCode' in bom_data and '|' in bom_data['BomPartCode']:
            try:
                part_no, part_code, description = [x.strip() for x in bom_data['BomPartCode'].split('|', 2)]
                bom_data['BomPartCode'] = part_no         # Store only part_no
                bom_data['BomPartDesc'] = description     # ✅ Store description
            except ValueError:
                return Response({'error': 'Invalid BomPartCode format. Use: "FG1001 | PC-001 | Description"'}, status=400)


        serializer = BOMItemSerializer(data=bom_data)
        if serializer.is_valid():
            serializer.save()
            return Response({'message': 'BOM item created successfully'}, status=201)

        return Response(serializer.errors, status=400)



from rest_framework.views import APIView
from rest_framework.response import Response
from django.db.models import Q
from .models import ItemTable
from .serializers import ItemWithBOMSerializer, ItemBasicSerializer

class ItemSearchWithBOMAPIView(APIView):
    def get(self, request):
        query = request.query_params.get('search', '')

        if not query:
            return Response([])

        # Check if the search query matches 'RM' or 'FG'
        is_rm_or_fg_search = query.upper() in ['RM', 'FG']
        
        if is_rm_or_fg_search:
            # Fetch items that have 'RM' or 'FG' in `main_group` or other matching criteria
            items = ItemTable.objects.filter(
                Q(main_group__icontains=query) |
                Q(part_no__icontains=query) |
                Q(Part_Code__icontains=query) |
                Q(Name_Description__icontains=query)
            )
            # Serialize items with BOM data (nested BOM)
            serializer = ItemWithBOMSerializer(items, many=True)
        else:
            # Fetch items with basic fields when search term is not 'RM' or 'FG'
            items = ItemTable.objects.filter(
                Q(part_no__icontains=query) |
                Q(Part_Code__icontains=query) |
                Q(Name_Description__icontains=query) |
                Q(main_group__icontains=query) |
                Q(item_group__icontains=query)
            )
            # Serialize only basic fields
            serializer = ItemBasicSerializer(items, many=True)

        return Response(serializer.data)


from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django.db.models import Q
from .models import ItemTable
from .serializers import ItemSearchSerializer2

class ItemSearchView2(APIView):
    def get(self, request):
        query = request.query_params.get('q', None)

        if not query:
            return Response({"detail": "Query parameter `q` is required."}, status=status.HTTP_400_BAD_REQUEST)

        # Search using OR filters
        items = ItemTable.objects.filter(
            Q(part_no__icontains=query) |
            Q(Part_Code__icontains=query) |
            Q(Name_Description__icontains=query)
        )

        serializer = ItemSearchSerializer2(items, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)


#Bom Routing & Bill of Material(BOM):- Fetch Scrap Code

from rest_framework import generics
from rest_framework.filters import SearchFilter
from .models import ItemTable
from .serializers import ItemTableSerializer3

class ItemTableScrapListView(generics.ListAPIView):
    # queryset = ItemTable.objects.filter(main_group='Scrap')  # FIXED here
    queryset = ItemTable.objects.filter(main_group__iexact='scrap')  # FIXED here
    serializer_class = ItemTableSerializer3
    filter_backends = [SearchFilter]
    search_fields = ['Name_Description', 'part_no']


from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from .models import ItemTable, Item
from django.contrib.auth.models import User

class CreateItemAPIView(APIView):
    def post(self, request):
        data = request.data
        supplier_id = data.get("supplier_id")
        if not supplier_id:
            return Response({"error": "supplier_id is required"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            supplier = Item.objects.get(id=supplier_id)
        except Item.DoesNotExist:
            return Response({"error": "Supplier not found"}, status=status.HTTP_404_NOT_FOUND)

        item = ItemTable.objects.create(
            part_no=data.get("part_no"),
            Name_Description=data.get("ItemDescription"),
            Item_Size=data.get("ItemSize"),
            Rate=data.get("Rate"),
            Unit_Code=data.get("Unit"),
            Note=data.get("Particular"),
            Store_Location=data.get("MakeMillName"),
            Route=data.get("DeliveryDate")
        )
        return Response({"message": "Item created", "item_id": item.id}, status=status.HTTP_201_CREATED)


from rest_framework.views import APIView
from rest_framework.response import Response
from .models import ItemTable, TaxDetails, Item

class GetItemDetailsAPIView(APIView):
    def get(self, request, part_no):
        try:
            item = ItemTable.objects.get(part_no=part_no)
        except ItemTable.DoesNotExist:
            return Response({"error": "Item not found"}, status=404)

        gst = TaxDetails.objects.filter(HSN_SAC_Code=item.HSN_SAC_Code).first()

        cgst = gst.CGST if gst else ''
        sgst = gst.SGST if gst else ''
        igst = gst.IGST if gst else ''
        cess = gst.Cess if gst else ''

        gst_details = {
            "part_no": item.part_no,
            "Part_Code": item.Part_Code or "",
            "HSN_SAC_Code": item.HSN_SAC_Code or "",
            "Rate": item.Rate or "",
            "Qty": "100",
            "SubTotal": "15000.0",
            "Discount": "10.0",
            "Packing": "",
            "Transport": "",
            "Assvalue": "",
            "TotalAmount": "159.30",
            "DiscRate": "135.00",
            "CGST": cgst,
            "CGSTAmt": "0.00",
            "SGST": sgst,
            "SGSTAmt": "90",
            "IGST": igst,
            "IGSTAmt": "89",
            "Vat": "89",
            "VatAmt": "70",
            "Cess": cess,
            "CessAmt": "80",
            "Total": "159.30"
        }

        response = {
            "data": {
                "id": item.id,
                "part_no": item.part_no,
                "Part_Code": item.Part_Code or "",
                "ItemDescription": item.Name_Description or "",
                "HSN_SAC_Code": item.HSN_SAC_Code or "",
                "CGST": cgst,
                "SGST": sgst,
                "Rate": item.Rate or "",
                "Disc": "10.0",
                "Qty": "100",
                "Unit": item.Unit_Code or "",
                "Particular": item.Note or "",
                "Mill_Name": item.Store_Location or "",
                "DeliveryDt": item.Route or "",
                "Schedule_Line": [
                    {
                        "part_no": item.part_no,
                        "Part_Code": item.Part_Code or "",
                        "ItemDescription": item.Name_Description or "",
                        "Qty": "100"
                    }
                ],
                "GST_Details": gst_details
            }
        }
        return Response(response)


# PO GST CALUCLATION

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from .models import ItemTransaction, Item

class ItemTransactionCreateAPIView(APIView):
    def post(self, request):
        data = request.data
        try:
            supplier = Item.objects.get(id=data['supplier_id'])
        except Item.DoesNotExist:
            return Response({"error": "Supplier not found"}, status=404)

        item_txn = ItemTransaction.objects.create(
            supplier=supplier,
            part_no=data.get('part_no'),
            ItemDescription=data.get('ItemDescription'),
            ItemSize=data.get('ItemSize'),
            Rate=data.get('Rate'),
            Disc=data.get('Disc'),
            Qty=data.get('Qty'),
            Unit=data.get('Unit'),
            Particular=data.get('Particular'),
            MakeMillName=data.get('MakeMillName'),
            DeliveryDate=data.get('DeliveryDate'),
            PartCode=data.get('PartCode')
        )
        return Response({"message": "Item transaction created", "id": item_txn.id}, status=201)



from rest_framework.response import Response
from rest_framework.views import APIView
from .models import ItemTransaction, ItemTable, TaxDetails


class ItemTransactionDetailAPIView(APIView):
    def get(self, request, pk):
        try:
            txn = ItemTransaction.objects.get(id=pk)
        except ItemTransaction.DoesNotExist:
            return Response({"error": "Transaction not found"}, status=404)

        item_info = ItemTable.objects.filter(part_no=txn.part_no).first()
        tax_info = TaxDetails.objects.filter(HSN_SAC_Code=item_info.HSN_SAC_Code).first() if item_info else None
        supplier = txn.supplier
        gst_code = supplier.GST_Tax_Code.strip().upper() if supplier and supplier.GST_Tax_Code else ""

        # Basic values
        rate = float(txn.Rate or 0)
        qty = float(txn.Qty or 0)
        subtotal = rate * qty
        discount_pct = float(txn.Disc or 0)
        discount_amt = round((subtotal * discount_pct) / 100, 2)
        ass_value = round(subtotal - discount_amt, 2)

        # Tax rates
        cgst_rate = float(tax_info.CGST or 0) if "CGST" in gst_code else 0
        sgst_rate = float(tax_info.SGST or 0) if "SGST" in gst_code else 0
        igst_rate = float(tax_info.IGST or 0) if "IGST" in gst_code else 0
        utgst_rate = 0  # Always 0 as per your clarification
        cess_rate = float(tax_info.Cess or 0) if "CESS" in gst_code else 0

        # Tax amounts
        cgst_amt = round((ass_value * cgst_rate) / 100, 2) if cgst_rate else ""
        sgst_amt = round((ass_value * sgst_rate) / 100, 2) if sgst_rate else ""
        igst_amt = round((ass_value * igst_rate) / 100, 2) if igst_rate else ""
        utgst_amt = ""  # Blank as UTGST is not used
        cess_amt = round((ass_value * cess_rate) / 100, 2) if cess_rate else ""

        total_amount = round(
            ass_value +
            (cgst_amt if cgst_amt != "" else 0) +
            (sgst_amt if sgst_amt != "" else 0) +
            (igst_amt if igst_amt != "" else 0) +
            (cess_amt if cess_amt != "" else 0),
            2
        )

        gst_details = {
            "part_no": txn.part_no,
            "Part_Code": item_info.Part_Code if item_info else "",
            "HSN_SAC_Code": item_info.HSN_SAC_Code if item_info else "",
            "Rate": rate,
            "Qty": qty,
            "SubTotal": subtotal,
            "Discount": discount_pct,
            "DiscountAmt": discount_amt,
            "Packing": "",
            "Transport": "",
            "Assvalue": ass_value,
            "CGST": cgst_rate if cgst_rate else "",
            "CGSTAmt": cgst_amt,
            "SGST": sgst_rate if sgst_rate else "",
            "SGSTAmt": sgst_amt,
            "IGST": igst_rate if igst_rate else "",
            "IGSTAmt": igst_amt,
            "UTGST": "",  # Set as blank
            "UTGSTAmt": "",  # Set as blank
            "Cess": cess_rate if cess_rate else "",
            "CessAmt": cess_amt,
            "Vat": "",
            "VatAmt": "",
            "Total": total_amount
        }

        response = {
            "data": {
                "id": txn.id,
                "part_no": txn.part_no,
                "Part_Code": item_info.Part_Code if item_info else "",
                "ItemDescription": txn.ItemDescription,
                "HSN_SAC_Code": item_info.HSN_SAC_Code if item_info else "",
                "CGST": gst_details["CGST"],
                "SGST": gst_details["SGST"],
                "IGST": gst_details["IGST"],
                "UTGST": gst_details["UTGST"],
                "Rate": rate,
                "Disc": discount_pct,
                "Qty": qty,
                "Unit": txn.Unit,
                "Particular": txn.Particular,
                "Mill_Name": txn.MakeMillName,
                "DeliveryDt": txn.DeliveryDate,
                "PartCode": txn.PartCode,
                "Schedule_Line": [
                    {
                        "part_no": txn.part_no,
                        "Part_Code": item_info.Part_Code if item_info else "",
                        "ItemDescription": txn.ItemDescription,
                        "Qty": qty
                    }
                ],
                "GST_Details": gst_details
            }
        }

        return Response(response)


# Item Search Api With BOM
class ItemSearchView5(APIView):
    def get(self, request):
        item_id = request.query_params.get('id', None)
        query = request.query_params.get('q', None)

        # Case 1: If id is provided → return single item
        if item_id:
            try:
                item = ItemTable.objects.get(id=item_id)
                serializer = ItemSearchSerializer5(item)
                return Response(serializer.data, status=status.HTTP_200_OK)
            except ItemTable.DoesNotExist:
                return Response({"detail": "Item not found."}, status=status.HTTP_404_NOT_FOUND)

        # Case 2: If query is provided → return search results
        elif query:
            items = ItemTable.objects.filter(
                Q(part_no__icontains=query) |
                Q(Part_Code__icontains=query) |
                Q(Name_Description__icontains=query)
            )
            serializer = ItemSearchSerializer5(items, many=True)
            return Response(serializer.data, status=status.HTTP_200_OK)

        # Case 3: No parameter → return all items
        else:
            items = ItemTable.objects.all()
            serializer = ItemSearchSerializer5(items, many=True)
            return Response(serializer.data, status=status.HTTP_200_OK)



# New Job Work Purchase Order
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from .models import ItemTransaction, Item

class ItemTransactionCreateAPIView2(APIView):
    def post(self, request):
        data = request.data
        try:
            supplier = Item.objects.get(id=data['supplier_id'])
        except Item.DoesNotExist:
            return Response({"error": "Supplier not found"}, status=404)

        item_txn = ItemTransaction2.objects.create(
            supplier=supplier,
            part_no=data.get('part_no'),
            ItemDescription=data.get('ItemDescription'),
            PartCode=data.get('PartCode'),
            Out=data.get('Out'),
            Inn=data.get('Inn'),
            Rate=data.get('Rate'),
            RType=data.get('RType'),
            Disc=data.get('Disc'),
            PoQty=data.get('PoQty'),
            Unit=data.get('Unit'),
            ParticularProcess=data.get('ParticularProcess')
        )
        return Response({"message": "Item transaction created", "id": item_txn.id}, status=201)




from rest_framework.response import Response
from rest_framework.views import APIView
from .models import ItemTransaction2, ItemTable, TaxDetails

class ItemTransactionDetailAPIView2(APIView):
    def get(self, request, pk):
        try:
            txn = ItemTransaction2.objects.get(id=pk)
        except ItemTransaction2.DoesNotExist:
            return Response({"error": "Transaction not found"}, status=404)

        item_info = ItemTable.objects.filter(part_no=txn.part_no).first()
        tax_info = TaxDetails.objects.filter(HSN_SAC_Code=item_info.SAC_Code).first() if item_info else None
        supplier = txn.supplier
        gst_code = supplier.GST_Tax_Code.strip().upper() if supplier and supplier.GST_Tax_Code else ""

        # Basic values
        rate = float(txn.Rate or 0)
        qty = float(txn.PoQty or 0)
        subtotal = round(rate * qty, 2)
        discount_pct = float(txn.Disc or 0)
        discount_amt = round((subtotal * discount_pct) / 100, 2)
        ass_value = round(subtotal - discount_amt, 2)

        # Tax rates
        if tax_info:
            cgst_rate = float(tax_info.CGST or 0) if "CGST" in gst_code else 0
            sgst_rate = float(tax_info.SGST or 0) if "SGST" in gst_code else 0
            igst_rate = float(tax_info.IGST or 0) if "IGST" in gst_code else 0
            utgst_rate = float(tax_info.UTGST or 0) if "UTGST" in gst_code else 0
        else:
            cgst_rate = sgst_rate = igst_rate = utgst_rate = 0

        # Tax amounts
        cgst_amt = round((ass_value * cgst_rate) / 100, 2) if cgst_rate else ""
        sgst_amt = round((ass_value * sgst_rate) / 100, 2) if sgst_rate else ""
        igst_amt = round((ass_value * igst_rate) / 100, 2) if igst_rate else ""
        utgst_amt = round((ass_value * utgst_rate) / 100, 2) if utgst_rate else ""

        # Total amount
        total_amount = round(
            ass_value +
            (cgst_amt if cgst_amt != "" else 0) +
            (sgst_amt if sgst_amt != "" else 0) +
            (igst_amt if igst_amt != "" else 0) +
            (utgst_amt if utgst_amt != "" else 0),
            2
        )

        gst_details = {
            "part_no": txn.part_no,
            "Part_Code": item_info.Part_Code if item_info else "",
            "SAC": item_info.SAC_Code if item_info else "",
            "Rate": rate,
            "Qty": qty,
            "SubTotal": subtotal,
            "Discount": discount_pct,
            "DiscountAmt": discount_amt,
            "Packing": "",
            "Transport": "",
            "Assvalue": ass_value,
            "CGST": cgst_rate if cgst_rate else "",
            "CGSTAmt": cgst_amt,
            "SGST": sgst_rate if sgst_rate else "",
            "SGSTAmt": sgst_amt,
            "IGST": igst_rate if igst_rate else "",
            "IGSTAmt": igst_amt,
            "UTGST": utgst_rate if utgst_rate else "",
            "UTGSTAmt": utgst_amt,
            "Total": total_amount
        }

        response = {
            "data": {
                "id": txn.id,
                "part_no": txn.part_no,
                "Part_Code": item_info.Part_Code if item_info else "",
                "ItemDescription": txn.ItemDescription,
                "SAC": item_info.SAC_Code if item_info else "",
                "CGST": gst_details["CGST"],
                "SGST": gst_details["SGST"],
                "IGST": gst_details["IGST"],
                "UTGST": gst_details["UTGST"],  # Added UTGST to main data block
                "Out": txn.Out,
                "Inn": txn.Inn,
                "Rate": rate,
                "RType": txn.RType,
                "Disc": discount_pct,
                "PoQty": txn.PoQty,
                "Unit": txn.Unit,
                "Particular": getattr(txn, 'ParticularProcess', ''),
                "Schedule_Line": [
                    {
                        "Part_Code": item_info.Part_Code if item_info else "",
                        "ItemDescription": txn.ItemDescription,
                        "PoQty": txn.PoQty,
                    }
                ],
                "GST_Details": gst_details
            }
        }

        return Response(response)


# BOM:- Item Part Master
class CombinedPartNoAPIView(APIView):
    def get(self, request):
        part_no = request.GET.get("part_no")
        operation_name = request.GET.get("operation_name")

        try:
            operation = Operation_Master_Model.objects.get(Operation_Name=operation_name)
        except Operation_Master_Model.DoesNotExist:
            return Response({"error": "Invalid Operation Name"}, status=400)

        if not part_no:
            return Response({"error": "part_no required"}, status=400)

        combined = f"{operation.Prefix}{part_no}"
        return Response({"combined_part_no": combined})


# RM Search Api
from rest_framework import generics, filters
from django.db.models import Q
from .models import ItemTable
from .serializers import RMItemTableSerializer

class RMItemSearchView(generics.ListAPIView):
    serializer_class = RMItemTableSerializer
    filter_backends = [filters.SearchFilter]
    search_fields = ['part_no', 'Part_Code', 'Name_Description']

    def get_queryset(self):
        return ItemTable.objects.filter(
            Q(main_group__iexact='RM') |
            Q(main_group__iexact='raw material')
        )



from rest_framework import generics
from .models import BOM_ItemPartMaster
from .serializers import BOMDropdownSerializer

class BOMDropdownAPIView(generics.ListAPIView):
    serializer_class = BOMDropdownSerializer

    def get_queryset(self):
        item_id = self.kwargs.get('item_id')
        return BOM_ItemPartMaster.objects.filter(item_id=item_id)


from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django.db.models import Q
from .models import BOMItem

class BOMItemSimpleSearch(APIView):
    def get(self, request, item_id, *args, **kwargs):
        search = request.query_params.get('Search', '').strip()

        # Filter by item_id
        items = BOMItem.objects.filter(item_id=item_id)

        # Optional search by OPNo or PartCode
        if search:
            items = items.filter(
                Q(OPNo__iexact=search) |
                Q(PartCode__iexact=search)
            )

        # Prepare the response
        data = []
        for item in items:
            data.append({
                "OPNo": item.OPNo,
                "Operation": "Cutting",  # Static; replace with dynamic if needed
                "PartCode": item.PartCode
            })

        return Response(data, status=status.HTTP_200_OK)
